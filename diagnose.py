import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import datetime
import unicodedata
from collections import defaultdict

def normalize(s):
    if not s: return ''
    s = ''.join(
        c for c in unicodedata.normalize('NFD', str(s).upper().strip())
        if unicodedata.category(c) != 'Mn'
    )
    s = s.replace('CC', 'C')
    return s

def val_to_sec(v):
    if v is None: return 0
    if isinstance(v, datetime.timedelta): return int(v.total_seconds())
    if isinstance(v, datetime.time): return v.hour*3600 + v.minute*60 + v.second
    if isinstance(v, datetime.datetime): return v.hour*3600 + v.minute*60 + v.second
    if isinstance(v, float):
        import math
        if math.isnan(v): return 0
        return int(round(v * 86400))
    if isinstance(v, str):
        parts = v.strip().split(':')
        if len(parts) >= 2:
            try:
                h = int(parts[0])
                m = int(parts[1])
                s = int(parts[2]) if len(parts)>2 else 0
                return h*3600 + m*60 + s
            except: pass
    return 0

def sec_to_str(sec):
    neg = sec < 0
    sec = abs(int(sec))
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return ('-' if neg else '') + f'{h:02d}:{m:02d}:{s:02d}'

CORTE = datetime.datetime(2026, 3, 16)

wb = openpyxl.load_workbook(
    r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\TEMPOS 2026.xlsx',
    data_only=True
)

# --- BASE analysis ---
ws = wb['BASE']
headers = [str(c.value) for c in ws[1]]
print("=== HEADERS ===")
for i, h in enumerate(headers):
    print(f"  col {i}: {h}")

# Map columns
COL_DATA = 0
COL_AGENTE = 3
COL_PAUSAS_TOTAL = 5
COL_TEMPO_LOGADO = 8
COL_META = 9
COL_CREDITO = 10
COL_DEFICIT = 11
COL_PAUSA = 12

# Aggregate by operator (from CORTE onwards)
ops = defaultdict(lambda: {
    'credito': 0, 'deficit': 0,
    'tempo_total': 0, 'meta_total': 0,
    'pausa_sum': 0.0, 'pausa_cnt': 0,
    'pausas_total_sum': 0, 'pausas_total_cnt': 0,
    'dias': 0
})

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    data_val = row[COL_DATA].value
    if not data_val: continue
    if isinstance(data_val, datetime.datetime):
        if data_val < CORTE: continue
    
    agente = str(row[COL_AGENTE].value).strip() if row[COL_AGENTE].value else ''
    if not agente: continue
    n_norm = normalize(agente)
    
    tempo = val_to_sec(row[COL_TEMPO_LOGADO].value)
    meta = val_to_sec(row[COL_META].value) or (7*3600 + 12*60)
    
    cred_val = row[COL_CREDITO].value
    def_val = row[COL_DEFICIT].value
    
    cred = val_to_sec(cred_val) if cred_val is not None else 0
    def_ = val_to_sec(def_val) if def_val is not None else 0
    
    # If both empty but we have tempo data, calculate
    if cred_val is None and def_val is None and tempo > 0:
        cred = max(0, tempo - meta)
        def_ = max(0, meta - tempo)
    
    pausa_val = row[COL_PAUSA].value
    pausa_total_val = row[COL_PAUSAS_TOTAL].value
    
    d = ops[n_norm]
    d['credito'] += cred
    d['deficit'] += def_
    d['tempo_total'] += tempo
    d['meta_total'] += meta
    d['dias'] += 1
    d['nome'] = agente
    
    if pausa_val is not None:
        try:
            p = float(pausa_val)
            if not (p != p):  # not NaN
                if 0 < abs(p) <= 1.0:
                    p *= 100
                d['pausa_sum'] += p
                d['pausa_cnt'] += 1
        except: pass
    
    if pausa_total_val is not None:
        pt = val_to_sec(pausa_total_val)
        d['pausas_total_sum'] += pt
        d['pausas_total_cnt'] += 1

# --- DATA COMPENSA ---
ws_comp = wb['DATA COMPENSA']
compensa = defaultdict(list)
for row in ws_comp.iter_rows(min_row=2, max_row=ws_comp.max_row):
    nome = str(row[0].value).strip() if row[0].value else ''
    data = row[1].value
    if nome and data is not None:
        n_norm = normalize(nome)
        compensa[n_norm].append(data)

print("\n=== OPERATOR SUMMARY (from CORTE) ===")
print(f"{'NOME':<35} {'DIAS':>4} {'CREDITO':>10} {'DEFICIT':>10} {'SALDO':>10} {'MEDIA_PAUSA':>11} {'COMPENSA':>8} {'SALDO_FINAL':>12}")
print("-" * 115)

for n_norm in sorted(ops.keys()):
    d = ops[n_norm]
    saldo = d['credito'] - d['deficit']
    media_pausa = d['pausa_sum'] / d['pausa_cnt'] if d['pausa_cnt'] > 0 else 0
    
    comp_count = len(compensa.get(n_norm, []))
    comp_sec = comp_count * (7*3600 + 12*60)  # 07:12:00 each
    
    saldo_final = saldo - comp_sec
    
    media_tempo = d['tempo_total'] / d['dias'] if d['dias'] > 0 else 0
    
    print(f"{d['nome'][:35]:<35} {d['dias']:>4} {sec_to_str(d['credito']):>10} {sec_to_str(d['deficit']):>10} {sec_to_str(saldo):>10} {media_pausa:>10.2f}% {comp_count:>8} {sec_to_str(saldo_final):>12}")

print("\n=== DATA COMPENSA DETAILS ===")
for n_norm in sorted(compensa.keys()):
    dates = compensa[n_norm]
    print(f"  {n_norm}: {len(dates)} dates -> {[str(d) for d in dates]}")

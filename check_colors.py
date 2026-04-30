import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

ADM = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\ADM EQUIPES.xlsx'

wb = openpyxl.load_workbook(ADM)
ws = wb['ESPELHO']

print("Verificando cores de fonte na coluna AGENTE (col A):")
print(f"{'Linha':<6} {'Nome':<45} {'Cor Fonte':<20} {'Vermelho?'}")
print("-"*90)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell = row[0]  # Coluna A = Agente
    nome = cell.value
    if not nome:
        continue
    font = cell.font
    cor = font.color if font else None
    cor_rgb = None
    is_red = False
    if cor:
        try:
            if cor.type == 'rgb':
                cor_rgb = cor.rgb
                # Vermelho puro ou variações: começa com FF e tem tons de vermelho
                r = int(cor_rgb[2:4], 16)
                g = int(cor_rgb[4:6], 16)
                b = int(cor_rgb[6:8], 16)
                # vermelho = R alto, G e B baixos
                if r > 150 and g < 100 and b < 100:
                    is_red = True
        except:
            pass
    print(f"{row[0].row:<6} {str(nome):<45} {str(cor_rgb):<20} {'⚠️ SIM' if is_red else 'não'}")

import openpyxl

FILES = {
    'META': r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\META GERAL 2026.xlsx'
}

try:
    wb = openpyxl.load_workbook(FILES['META'], data_only=True)
    ws = wb['METAS MAIO2026']
    
    # Header is in row 4 (index 3)
    # Col 1: Agente, Col 4: H.O, Col 5: COMISSÃO
    print("Row 4 (Headers):", [c.value for c in ws[4]])
    
    print("\nData rows:")
    for row in ws.iter_rows(min_row=5, max_row=15):
        agente = row[1].value
        ho = row[3].value
        comissao = row[4].value
        print(f"Agente: {agente} | H.O: {ho} | COMISSÃO: {comissao}")
        
except Exception as e:
    print(f"Error: {e}")

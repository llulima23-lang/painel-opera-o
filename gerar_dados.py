import os
import sys
import json
import openpyxl
import pandas as pd
import datetime
import unicodedata
import traceback
import time

try:
    sys.stdout.reconfigure(encoding='utf-8')
except:
    pass

def get_path(path_key, default_path):
    if os.path.exists(default_path):
        return default_path
    base_name = os.path.basename(default_path)
    local_candidate = os.path.join(os.path.dirname(os.path.abspath(__file__)), base_name)
    if os.path.exists(local_candidate):
        return local_candidate
    local_candidate_sibling = os.path.join(os.getcwd(), base_name)
    if os.path.exists(local_candidate_sibling):
        return local_candidate_sibling
    return default_path

FILES = {
    'META': get_path('META', r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\META GERAL 2026.xlsx'),
    'TEMPOS': get_path('TEMPOS', r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\TEMPOS 2026.xlsx'),
    'ADM': get_path('ADM', r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\ADM EQUIPES.xlsx'),
    'ABS': get_path('ABS', r'C:\Users\sup.luciana\Meu Drive\MF\MF\ABS\TRATADO - ABS.xlsx')
}
JS_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data_embedded.js')

CORTE_BH = datetime.date(2026, 3, 16)
META_PADRAO_SEC = 7 * 3600 + 12 * 60  # 07:12:00
COMPENSA_SEC    = 7 * 3600 + 12 * 60  # 07:12:00 por data compensada



def normalize(s):
    """Remove acentos e converte para maiúsculas para comparação."""
    if not s: return ''
    s = ''.join(
        c for c in unicodedata.normalize('NFD', str(s).upper().strip())
        if unicodedata.category(c) != 'Mn'
    )
    # Remove espaços duplos
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s

def fuzzy_match(name_a, name_b):
    """Verifica se dois nomes normalizados se referem à mesma pessoa.
    Usa lógica: um nome é prefixo do outro, ou compartilham primeiros 2+ termos."""
    if name_a == name_b:
        return True
    if not name_a or not name_b:
        return False
    # Um começa com o outro
    if name_a.startswith(name_b) or name_b.startswith(name_a):
        return True
    # Compara primeiros 2 termos (primeiro nome + sobrenome)
    parts_a = name_a.split()
    parts_b = name_b.split()
    if len(parts_a) >= 2 and len(parts_b) >= 2:
        # Pelo menos primeiro nome + segundo nome iguais e terceiro (se existir)
        if parts_a[0] == parts_b[0] and parts_a[1] == parts_b[1]:
            # Se tem 3+ termos, verificar o terceiro também para evitar falsos positivos
            if len(parts_a) >= 3 and len(parts_b) >= 3:
                return parts_a[2] == parts_b[2]
            return True
    return False

def find_best_match(target_norm, name_set):
    """Encontra o melhor match para um nome normalizado em um set de nomes."""
    if target_norm in name_set:
        return target_norm
    for candidate in name_set:
        if fuzzy_match(target_norm, candidate):
            return candidate
    return None

def val_to_sec(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return 0
    if isinstance(v, datetime.timedelta): return int(v.total_seconds())
    if isinstance(v, datetime.time):      return v.hour*3600 + v.minute*60 + v.second
    if isinstance(v, datetime.datetime):  return v.hour*3600 + v.minute*60 + v.second
    if isinstance(v, float):              return int(round(v * 86400))
    if isinstance(v, str):
        parts = v.strip().split(':')
        if len(parts) >= 2:
            try: return int(parts[0])*3600 + int(parts[1])*60 + (int(parts[2]) if len(parts)>2 else 0)
            except: pass
    return 0

def safe_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    f = float(v)
    return round(f*100, 2) if 0 < abs(f) <= 1.0 else round(f, 2)

def is_red_font(cell):
    try:
        color = cell.font.color
        if not color: return False
        # Vermelho padrão Excel (RGB: FFFF0000 ou Indexed: 2)
        if color.type == 'rgb' and str(color.rgb).upper() in ('FFFF0000', 'FF0000'):
            return True
        if color.type == 'indexed' and color.indexed == 2:
            return True
        # Tenta extrair RGB se for outro formato
        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb).upper()
            if len(rgb) == 8:
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
                return r > 150 and g < 100 and b < 100
    except: pass
    return False

def process_excel():
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Processando planilhas...")

    # ── 1. ADM EQUIPES (cores de fonte para detectar desligados) ──────
    for attempt in range(3):
        try:
            wb_adm  = openpyxl.load_workbook(FILES['ADM'], data_only=True)
            ws_esp  = wb_adm['ESPELHO']
            break
        except Exception as e:
            if attempt == 2: raise e
            print(f"  Aviso: Erro ao abrir ADM (tentativa {attempt+1}): {e}")
            time.sleep(2)
    
    header_map = {str(c.value).strip(): c.column-1 for c in ws_esp[1] if c.value}

    col_nome = header_map.get('Agente',              0)
    col_adm  = header_map.get('Admissão',            3)
    col_mat  = header_map.get('Matricula',            5)
    col_op   = header_map.get('Operação que atua',   16)

    adm_list         = []   # operadores ativos
    desligados_norm  = set()
    operacoes_set    = set()
    adm_norm_set     = set()  # Nomes normalizados dos operadores ativos

    for row in ws_esp.iter_rows(min_row=2, max_row=ws_esp.max_row):
        nome_cell = row[col_nome]
        nome = nome_cell.value
        if not nome: continue
        nome = str(nome).strip()

        if is_red_font(nome_cell):
            desligados_norm.add(normalize(nome))
            continue

        adm_val = row[col_adm].value if col_adm < len(row) else None
        mat_val = row[col_mat].value if col_mat < len(row) else None
        op_val  = row[col_op].value  if col_op  < len(row) else None

        adm_str = (adm_val.strftime('%Y-%m-%d') if isinstance(adm_val, (datetime.datetime, datetime.date))
                   else str(adm_val) if adm_val else None)
        mat_str = None
        if mat_val is not None:
            try:   mat_str = str(int(float(mat_val)))
            except: mat_str = str(mat_val)

        op_str = str(op_val).strip() if op_val else None
        if op_str: operacoes_set.add(op_str)

        nome_norm = normalize(nome)
        adm_norm_set.add(nome_norm)

        adm_list.append({
            'nome':      nome,
            'nome_norm': nome_norm,
            'admissao':  adm_str,
            'matricula': mat_str,
            'operacao':  op_str,
            'foto':      None
        })

    print(f"  ADM: {len(adm_list)} ativos | {len(desligados_norm)} desligados excluidos")
    print(f"  Operacoes: {sorted(operacoes_set)}")

    # ── 1.5 MAPEAMENTO DE FOTOS ──────────────────────────────────────
    import shutil, glob
    image_extensions = ('.jpg', '.jpeg', '.png')
    fotos_dir_dst = os.path.join(os.path.dirname(__file__), 'fotos')
    os.makedirs(fotos_dir_dst, exist_ok=True)

    fotos_dir_src_default = r'C:\Users\sup.luciana\Meu Drive\FOTOS EQUIPES'
    if os.path.exists(fotos_dir_src_default):
        fotos_dir_src = fotos_dir_src_default
    else:
        fotos_dir_src = os.path.join(os.path.dirname(__file__), 'fotos_src')
        if not os.path.exists(fotos_dir_src):
            fotos_dir_src = fotos_dir_dst
            
    src_images = []
    if os.path.exists(fotos_dir_src):
        for root, dirs, files in os.walk(fotos_dir_src):
            for file in files:
                if file.lower().endswith(image_extensions):
                    src_images.append(os.path.join(root, file))

    fotos_found = 0
    for op in adm_list:
        n_norm = op['nome_norm']
        parts = n_norm.split()
        op_first = parts[0] if parts else ''
        best = None

        for img_path in src_images:
            base = normalize(os.path.splitext(os.path.basename(img_path))[0])
            if fuzzy_match(n_norm, base):
                best = img_path
                break
            if op_first and base.startswith(op_first + ' ') and len(parts) > 1 and parts[1] in base:
                best = img_path
                break

        if not best and op_first:
            for img_path in src_images:
                base = normalize(os.path.splitext(os.path.basename(img_path))[0])
                if base == op_first:
                    best = img_path
                    break

        if best:
            ext = os.path.splitext(best)[1].lower()
            safe = op['matricula'] if op['matricula'] else n_norm.replace(' ', '_')
            dst = os.path.join(fotos_dir_dst, f"{safe}{ext}")
            try:
                if os.path.abspath(best) != os.path.abspath(dst):
                    shutil.copy2(best, dst)
                op['foto'] = f"fotos/{safe}{ext}"
                fotos_found += 1
            except Exception as e:
                print(f"  Erro foto {best}: {e}")

    print(f"  FOTOS: {fotos_found}/{len(adm_list)} encontradas")

    # ── 2. TEMPOS BASE agrupado por operador × mes ────────────────────
    for attempt in range(3):
        try:
            df_base = pd.read_excel(FILES['TEMPOS'], sheet_name='BASE', header=0)
            break
        except Exception as e:
            if attempt == 2: raise e
            print(f"  Aviso: Erro ao abrir TEMPOS (tentativa {attempt+1}): {e}")
            time.sleep(2)
    df_base['DATA'] = pd.to_datetime(df_base['DATA'], errors='coerce')
    df_base = df_base[df_base['DATA'] >= pd.Timestamp(CORTE_BH)].copy()
    df_base['_NORM'] = df_base['AGENTE'].apply(lambda x: normalize(str(x)))

    # Criar mapa de resolução de nomes: TEMPOS -> ADM
    # Para cada nome único no TEMPOS, encontrar o equivalente no ADM
    tempo_names = set(df_base['_NORM'].unique())
    name_resolve = {}  # tempo_norm -> adm_norm
    unmatched_tempos = set()

    for tn in tempo_names:
        if tn in desligados_norm:
            name_resolve[tn] = None  # Desligado, ignorar
            continue
        # Verifica se esse nome de desligado faz match fuzzy
        is_desligado = False
        for dn in desligados_norm:
            if fuzzy_match(tn, dn):
                is_desligado = True
                break
        if is_desligado:
            name_resolve[tn] = None
            continue

        match = find_best_match(tn, adm_norm_set)
        if match:
            name_resolve[tn] = match
        else:
            unmatched_tempos.add(tn)
            name_resolve[tn] = tn  # Usa o próprio nome

    if unmatched_tempos:
        print(f"  ⚠️  Nomes TEMPOS sem match no ADM: {sorted(unmatched_tempos)}")

    # Filtra desligados
    df_base = df_base[df_base['_NORM'].apply(lambda x: name_resolve.get(x) is not None)]
    
    # Resolve nomes para o padrão ADM
    df_base['_NORM_RESOLVED'] = df_base['_NORM'].apply(lambda x: name_resolve.get(x, x))
    df_base['_MES'] = df_base['DATA'].dt.strftime('%Y-%m')
    print(f"  TEMPOS: {len(df_base)} registros a partir de {CORTE_BH}")

    # Detectar nomes de colunas (podem ter variações de acentuação)
    col_credito = None
    col_deficit = None
    col_pausa = None
    col_pausas_total = None
    col_tempo_logado = None
    col_meta = None

    for c in df_base.columns:
        cn = normalize(str(c))
        if cn == 'CREDITO': col_credito = c
        elif cn == 'DEFICT' or cn == 'DEFICIT': col_deficit = c
        elif cn == 'PAUSA' and 'TOTAL' not in cn and 'PAUSAS' not in cn: col_pausa = c
        elif 'PAUSAS TOTAL' in cn or cn == 'PAUSAS TOTAL': col_pausas_total = c
        elif cn == 'TEMPO LOGADO': col_tempo_logado = c
        elif cn == 'META': col_meta = c

    # Fallback names
    if not col_credito: col_credito = 'Crédito'
    if not col_deficit: col_deficit = 'DEFICT'
    if not col_pausa: col_pausa = 'Pausa'
    if not col_pausas_total: col_pausas_total = 'Pausas total'
    if not col_tempo_logado: col_tempo_logado = 'TEMPO LOGADO'
    if not col_meta: col_meta = 'META'

    print(f"  Colunas: credito='{col_credito}', deficit='{col_deficit}', pausa='{col_pausa}', pausas_total='{col_pausas_total}'")

    # agrupamento (nome_norm_resolved, mes) → stats
    monthly_map = {}
    for _, row in df_base.iterrows():
        nome_norm = row['_NORM_RESOLVED']
        mes       = row['_MES']
        key       = (nome_norm, mes)

        tempo_log = val_to_sec(row.get(col_tempo_logado, 0))
        meta_sec  = val_to_sec(row.get(col_meta, None)) or META_PADRAO_SEC

        # Cálculo manual conforme solicitado pela usuária:
        # A diferença entre Tempo Logado e Meta define o crédito ou débito do dia.
        diff = tempo_log - meta_sec
        if diff > 0:
            credito_sec = diff
            deficit_sec = 0
        else:
            credito_sec = 0
            deficit_sec = abs(diff)

        # Pausa % (coluna "Pausa" - é uma fração decimal como 0.123 = 12.3%)
        pausa_pct_val = row.get(col_pausa, None)

        # Pausas total (coluna "Pausas total" - é HH:MM:SS)
        pausas_total_val = row.get(col_pausas_total, None)

        if key not in monthly_map:
            monthly_map[key] = {
                'nome_norm': nome_norm, 'mes': mes,
                'credito': 0, 'deficit': 0, 'dias': 0,
                'total_tempo': 0, 'total_meta': 0,
                'pausa_pct_sum': 0.0, 'pausa_pct_dias': 0,
                'pausas_total_sec': 0, 'pausas_total_dias': 0
            }
        d = monthly_map[key]
        d['credito']    += credito_sec
        d['deficit']    += deficit_sec
        d['dias']       += 1
        d['total_tempo']+= tempo_log
        d['total_meta'] += meta_sec

        # Pausa % - lê e converte
        if pausa_pct_val is not None and not (isinstance(pausa_pct_val, float) and pd.isna(pausa_pct_val)):
            p = float(pausa_pct_val)
            if 0 < abs(p) <= 1.0: p *= 100
            d['pausa_pct_sum'] += p
            d['pausa_pct_dias'] += 1

        # Pausas total (absoluto em segundos)
        if pausas_total_val is not None and not (isinstance(pausas_total_val, float) and pd.isna(pausas_total_val)):
            pt = val_to_sec(pausas_total_val)
            d['pausas_total_sec'] += pt
            d['pausas_total_dias'] += 1

    # Converte para lista com médias calculadas
    tempos_list = []
    # Precisamos do nome "amigável" - pegar do ADM
    adm_name_map = {a['nome_norm']: a['nome'] for a in adm_list}

    for (nome_norm, mes), v in monthly_map.items():
        dias = v['dias'] or 1
        media_tempo = int(v['total_tempo'] / dias)
        media_meta  = int(v['total_meta']  / dias)
        media_pausa_pct = round((v['pausas_total_sec'] / v['total_tempo']) * 100, 2) if v['total_tempo'] > 0 else 0
        media_pausas_total_sec = int(v['pausas_total_sec'] / v['pausas_total_dias']) if v['pausas_total_dias'] > 0 else 0

        nome_display = adm_name_map.get(nome_norm, nome_norm)

        tempos_list.append({
            'nome':             nome_display,
            'nome_norm':        nome_norm,
            'mes':              mes,                     # "2026-03"
            'credito_sec':      v['credito'],
            'deficit_sec':      v['deficit'],
            'dias_trabalhados': v['dias'],
            'media_tempo_sec':  media_tempo,
            'meta_diaria_sec':  media_meta,
            'media_pausa_pct':  media_pausa_pct,
            'media_pausas_total_sec': media_pausas_total_sec,
        })

    print(f"  TEMPOS: {len(tempos_list)} registros mes/operador")

    # ── 2.5 DATA COMPENSA ─────────────────────────────────────────────
    compensa_map = {}  # adm_norm -> list of ISO dates
    try:
        df_comp = pd.read_excel(FILES['TEMPOS'], sheet_name='DATA COMPENSA', header=0)
        for _, row in df_comp.iterrows():
            nome_c = str(row.get('NOME', '')).strip()
            data_c = row.get('DATA', None)
            if not nome_c:
                continue

            n_norm_raw = normalize(nome_c)

            # Resolve para o nome ADM usando fuzzy match
            resolved = find_best_match(n_norm_raw, adm_norm_set)
            if not resolved:
                resolved = n_norm_raw
                print(f"  ⚠️  COMPENSA: '{nome_c}' (norm: {n_norm_raw}) sem match no ADM")

            if resolved not in compensa_map:
                compensa_map[resolved] = []

            # Só adiciona se tem data preenchida
            if pd.notna(data_c) and data_c is not None:
                if isinstance(data_c, (datetime.datetime, datetime.date)):
                    compensa_map[resolved].append(data_c.strftime('%Y-%m-%d'))
                elif isinstance(data_c, str) and data_c.strip():
                    compensa_map[resolved].append(data_c.strip())

        # Remove entradas sem datas
        compensa_map = {k: v for k, v in compensa_map.items() if v}
        print(f"  COMPENSA: {len(compensa_map)} operadores com datas de compensação")
        for k, v in compensa_map.items():
            print(f"    {k}: {v}")
    except Exception as e:
        print(f"  Erro ao ler DATA COMPENSA: {e}")

    # ── 2.7 RESUMO (Banco de Horas Inicial e Saldo Final com cor de fonte) ──
    resumo_map = {} # adm_norm -> bh_inicial_sec
    resumo_saldo_final = {} # adm_norm -> saldo final em sec (coluna F)
    resumo_credito_total = 0  # Soma total da coluna CRÉDITO
    resumo_debito_total = 0   # Soma total da coluna Débito final
    resumo_credito_individual = {}  # adm_norm -> credito_sec
    resumo_debito_individual = {}   # adm_norm -> debito_final_sec
    try:
        # Usamos openpyxl para detectar a COR da fonte (vermelho = negativo)
        for attempt in range(3):
            try:
                wb_res = openpyxl.load_workbook(FILES['TEMPOS'], data_only=True)
                ws_res = wb_res['RESUMO']
                break
            except Exception as e:
                if attempt == 2: raise e
                time.sleep(2)

        # Detecta colunas
        header_res = [str(c.value).upper() if c.value else '' for c in ws_res[1]]
        idx_nome = header_res.index('NOME') if 'NOME' in header_res else 0
        
        # Coluna CRÉDITO (B)
        idx_credito_res = -1
        for i, h in enumerate(header_res):
            hn = normalize(h)
            if hn == 'CREDITO':
                idx_credito_res = i
                break
        if idx_credito_res == -1: idx_credito_res = 1  # Fallback col B
        
        # Coluna Débito final (E)
        idx_debito_final = -1
        for i, h in enumerate(header_res):
            hn = normalize(h)
            if 'DEBITO FINAL' in hn:
                idx_debito_final = i
                break
        if idx_debito_final == -1: idx_debito_final = 4  # Fallback col E
        
        # Procura coluna de Saldo Final
        idx_saldo = -1
        for i, h in enumerate(header_res):
            if 'SALDO FINAL' == h:
                idx_saldo = i
                break
        if idx_saldo == -1:
            for i, h in enumerate(header_res):
                if 'SALDO FINAL' in h:
                    idx_saldo = i
                    break
        if idx_saldo == -1:
            for i, h in enumerate(header_res):
                if 'DEBITO FINAL' in h or 'DÉBITO FINAL' in h:
                    idx_saldo = i
                    break
        if idx_saldo == -1: idx_saldo = 5

        idx_bh_ini = 3 # Coluna D (banco de horas)

        for row in ws_res.iter_rows(min_row=2, max_row=ws_res.max_row):
            nome_res = str(row[idx_nome].value).strip() if row[idx_nome].value else None
            if not nome_res: continue
            
            n_norm = normalize(nome_res)
            resolved = find_best_match(n_norm, adm_norm_set)
            if not resolved: resolved = n_norm
            
            # Verifica se é um operador real (não nota de rodapé)
            is_operator = resolved in adm_norm_set
            
            # Banco de Horas Inicial (Coluna D)
            cell_bh = row[idx_bh_ini]
            val_bh = val_to_sec(cell_bh.value)
            if is_red_font(cell_bh): val_bh = -abs(val_bh)
            resumo_map[resolved] = val_bh
            
            # Saldo Final (Coluna F ou similar)
            cell_saldo = row[idx_saldo]
            val_saldo = val_to_sec(cell_saldo.value)
            if is_red_font(cell_saldo): 
                val_saldo = -abs(val_saldo)
            resumo_saldo_final[resolved] = val_saldo
            
            # Crédito e Débito Final (para totais do dashboard)
            if is_operator:
                cell_cred = row[idx_credito_res]
                val_cred = val_to_sec(cell_cred.value)
                resumo_credito_total += val_cred
                resumo_credito_individual[resolved] = val_cred
                
                cell_deb = row[idx_debito_final]
                val_deb = val_to_sec(cell_deb.value)
                resumo_debito_total += val_deb
                resumo_debito_individual[resolved] = val_deb

        print(f"  RESUMO: {len(resumo_map)} saldos | Crédito Total={resumo_credito_total}s | Débito Total={resumo_debito_total}s")
    except Exception as e:
        print(f"  Erro ao ler RESUMO: {e}")
        traceback.print_exc()

    # ── 3. META GERAL (Busca dinâmica de abas) ────────────────────────
    for attempt in range(3):
        try:
            xls_meta = pd.ExcelFile(FILES['META'])
            break
        except Exception as e:
            if attempt == 2: raise e
            print(f"  Aviso: Erro ao abrir META (tentativa {attempt+1}): {e}")
            time.sleep(2)
    # Pega todas as abas que começam com METAS, exceto possíveis backups
    meta_sheets = [s for s in xls_meta.sheet_names if s.startswith('METAS') and 'Backup' not in s]
    meta_data   = {}

    for sh in meta_sheets:
        df_raw = pd.read_excel(xls_meta, sheet_name=sh, header=None, nrows=4)
        du_val = None
        for r in range(df_raw.shape[0]):
            for c in range(df_raw.shape[1]-1):
                if str(df_raw.iloc[r,c]).strip() == 'D.U':
                    try: du_val = int(df_raw.iloc[r,c+1])
                    except: pass
        df = pd.read_excel(xls_meta, sheet_name=sh, header=3)
        rows = []
        for _, row in df.iterrows():
            nome = row.get('Agente', None)
            if not nome or str(nome).strip() in ('','nan','None'): continue
            nome_str  = str(nome).strip()
            nome_norm = normalize(nome_str)

            # Verifica desligados com fuzzy match
            is_desligado = False
            for dn in desligados_norm:
                if fuzzy_match(nome_norm, dn):
                    is_desligado = True
                    break
            if is_desligado:
                continue

            # Resolve nome para ADM
            resolved_norm = find_best_match(nome_norm, adm_norm_set)
            if resolved_norm:
                nome_norm = resolved_norm

            promessas = row.get('PROMESSAS', None)
            meta_prom = row.get('META PROM', None)
            qualidade = safe_pct(row.get('QUALIDADE', None))
            abs_col   = next((c for c in df.columns if str(c).upper().startswith('ABS')), None)
            abs_dias  = float(row[abs_col]) if abs_col and not pd.isna(row.get(abs_col)) else 0
            mat_val   = row.get('Matricula', None)
            mat_str   = None
            if mat_val is not None and not (isinstance(mat_val, float) and pd.isna(mat_val)):
                try:   mat_str = str(int(float(mat_val)))
                except: mat_str = str(mat_val)

            # Busca dinâmica de colunas para H.O e Comissão (robusto contra espaços/acentos)
            col_ho_name = next((c for c in df.columns if normalize(c) == 'HO'), 'H.O')
            col_meta_ho_name = next((c for c in df.columns if normalize(c) in ('META H.O', 'META HO')), 'META H.O')
            col_com_name = next((c for c in df.columns if normalize(c) == 'COMISSAO'), 'COMISSÃO')
            
            val_ho   = float(row.get(col_ho_name, 0)) if not pd.isna(row.get(col_ho_name)) else 0
            meta_ho_val = row.get(col_meta_ho_name, None)
            val_meta_ho = float(meta_ho_val) if meta_ho_val is not None and not (isinstance(meta_ho_val, float) and pd.isna(meta_ho_val)) else 0

            val_com  = float(row.get(col_com_name, 0)) if not pd.isna(row.get(col_com_name)) else 0

            # Pausas (percentual) e Banco de Horas (tempo)
            col_pausa_name = next((c for c in df.columns if normalize(c) == 'PAUSAS' or normalize(c) == 'PAUSA'), 'Pausas')
            pausas_val = row.get(col_pausa_name, 0)
            pausas_pct = float(pausas_val) * 100 if not pd.isna(pausas_val) else 0

            # Quartil e Dispersão (se existirem)
            col_quartil_name = next((c for c in df.columns if normalize(c) == 'QUARTIL'), None)
            col_dispersao_name = next((c for c in df.columns if normalize(c) in ('DISPERSAO', 'DISPERSÃO')), None)

            quartil_val = row.get(col_quartil_name) if col_quartil_name else None
            dispersao_val = row.get(col_dispersao_name) if col_dispersao_name else None

            quartil_str = str(quartil_val).strip() if pd.notna(quartil_val) and quartil_val is not None else None
            dispersao_num = safe_pct(dispersao_val) if pd.notna(dispersao_val) and dispersao_val is not None else None

            rows.append({
                'nome':      nome_str,
                'nome_norm': nome_norm,
                'matricula': mat_str,
                'promessas': float(promessas) if promessas is not None and not (isinstance(promessas,float) and pd.isna(promessas)) else 0,
                'meta_prom': float(meta_prom)  if meta_prom  is not None and not (isinstance(meta_prom, float) and pd.isna(meta_prom))  else 0,
                'qualidade': qualidade,
                'abs_dias':  abs_dias,
                'ho': val_ho,
                'meta_ho': val_meta_ho,
                'comissao': val_com,
                'pausas': pausas_pct,
                'banco_horas': val_to_sec(row.get('Banco de horas', 0)),
                'quartil': quartil_str,
                'dispersao': dispersao_num
            })
        meta_data[sh] = {'rows': rows, 'du': du_val}
        print(f"  META {sh}: {len(rows)} ativos, D.U={du_val}")

    # ── Debug: mostra saldos finais ───────────────────────────────────
    print("\n  === SALDOS BANCO DE HORAS ===")
    bh_totals = {}
    for t in tempos_list:
        nn = t['nome_norm']
        if nn not in bh_totals:
            bh_totals[nn] = {'nome': t['nome'], 'credito': 0, 'deficit': 0}
        bh_totals[nn]['credito'] += t['credito_sec']
        bh_totals[nn]['deficit'] += t['deficit_sec']

    def sec_fmt(s):
        neg = s < 0
        s = abs(int(s))
        h = s // 3600; m = (s % 3600) // 60; ss = s % 60
        return ('-' if neg else '') + f'{h:02d}:{m:02d}:{ss:02d}'

    for nn in sorted(bh_totals.keys()):
        bh = bh_totals[nn]
        comp_dates = compensa_map.get(nn, [])
        comp_sec = len(comp_dates) * COMPENSA_SEC
        saldo = bh['credito'] - bh['deficit'] - comp_sec
        print(f"    {bh['nome'][:35]:35} C={sec_fmt(bh['credito']):>10} D={sec_fmt(bh['deficit']):>10} Comp={len(comp_dates)}x07:12={sec_fmt(comp_sec):>10} SALDO={sec_fmt(saldo):>10}")

    # ── 4. MÉDIA TEMPO LOGADO MENSAL (global) ────────────────────────
    tempo_logado_media_mensal = {}
    meses_tempos = {}
    for t in tempos_list:
        mes = t['mes']
        if mes not in meses_tempos:
            meses_tempos[mes] = {'total_sec': 0, 'count': 0}
        meses_tempos[mes]['total_sec'] += t['media_tempo_sec']
        meses_tempos[mes]['count'] += 1
    for mes, v in meses_tempos.items():
        if v['count'] > 0:
            tempo_logado_media_mensal[mes] = int(v['total_sec'] / v['count'])
    print(f"  TEMPO LOGADO MÉDIAS: {tempo_logado_media_mensal}")

    # ── 5. ABSENTEÍSMO (TRATADO - ABS.xlsx) ───────────────────────────
    abs_data = {
        'geral_por_mes': {},        # mes_key -> abs_pct (linha 10)
        'por_operacao_mes': {},     # mes_key -> [{operacao, abs_pct}]
        'individual_mes': {},       # mes_key -> [{nome, nome_norm, total_dias, operacao}]
        'geral_ultimos_3': 0.0,     # média dos últimos 3 meses
    }
    try:
        wb_abs = openpyxl.load_workbook(FILES['ABS'], data_only=True)
        # Pega abas 2026 (últimas disponíveis)
        abs_sheets_2026 = [s for s in wb_abs.sheetnames if '2026' in s]
        print(f"  ABS: Abas 2026 encontradas: {abs_sheets_2026}")

        mes_name_map = {
            'JANEIRO': '2026-01', 'FEVEREIRO': '2026-02', 'MARCO': '2026-03', 'MARÇO': '2026-03',
            'ABRIL': '2026-04', 'MAIO': '2026-05', 'JUNHO': '2026-06', 'JULHO': '2026-07',
            'AGOSTO': '2026-08', 'SETEMBRO': '2026-09', 'OUTUBRO': '2026-10',
            'NOVEMBRO': '2026-11', 'DEZEMBRO': '2026-12'
        }

        for sn in abs_sheets_2026:
            ws_abs = wb_abs[sn]
            # Determinar mes_key a partir do nome da aba
            sn_upper = normalize(sn).replace('ABS ', '').replace('2026', '').strip()
            mes_key = mes_name_map.get(sn_upper, None)
            if not mes_key:
                for k, v in mes_name_map.items():
                    if k in normalize(sn):
                        mes_key = v
                        break
            if not mes_key:
                print(f"    ⚠️  ABS: Não consegui determinar mês para aba '{sn}'")
                continue

            # Linha 10: ABS geral (totalização da equipe)
            abs_geral_cell = ws_abs.cell(row=10, column=10)  # Coluna J (10)
            abs_geral_val = abs_geral_cell.value
            if abs_geral_val is not None:
                try:
                    abs_pct = float(abs_geral_val) * 100 if float(abs_geral_val) < 1 else float(abs_geral_val)
                except:
                    abs_pct = 0
            else:
                abs_pct = 0
            abs_data['geral_por_mes'][mes_key] = round(abs_pct, 2)

            # Linhas 6-9: ABS por operação
            ops_abs = []
            for r in range(6, 10):
                gestao_cell = ws_abs.cell(row=r, column=3)  # Col C = GESTÃO/PRODUTO
                abs_cell = ws_abs.cell(row=r, column=10)     # Col J = ABS
                gestao = str(gestao_cell.value).strip() if gestao_cell.value else None
                if not gestao or gestao == 'None': continue
                abs_op_val = abs_cell.value
                if abs_op_val is not None:
                    try:
                        abs_op_pct = float(abs_op_val) * 100 if float(abs_op_val) < 1 else float(abs_op_val)
                    except:
                        abs_op_pct = 0
                else:
                    abs_op_pct = 0
                ops_abs.append({'operacao': gestao, 'abs_pct': round(abs_op_pct, 2)})
            abs_data['por_operacao_mes'][mes_key] = ops_abs

            # Linhas 14+: ABS individual (OPERADOR + TOTAL DE DIAS)
            ind_abs = []
            for r in range(14, ws_abs.max_row + 1):
                op_cell = ws_abs.cell(row=r, column=1)   # Col A = OPERAÇÃO
                nome_cell = ws_abs.cell(row=r, column=2) # Col B = OPERADOR
                total_cell = ws_abs.cell(row=r, column=5) # Col E = TOTAL DE DIAS
                nome_val = str(nome_cell.value).strip() if nome_cell.value else None
                if not nome_val or nome_val == 'None': continue
                op_val = str(op_cell.value).strip() if op_cell.value else ''
                total_dias = 0
                if total_cell.value is not None:
                    try: total_dias = int(float(total_cell.value))
                    except: total_dias = 0
                nome_norm_abs = normalize(nome_val)
                resolved_abs = find_best_match(nome_norm_abs, adm_norm_set) or nome_norm_abs
                ind_abs.append({
                    'nome': nome_val,
                    'nome_norm': resolved_abs,
                    'total_dias': total_dias,
                    'operacao': op_val
                })
            abs_data['individual_mes'][mes_key] = ind_abs
            print(f"    ABS {sn}: geral={abs_pct:.2f}% | {len(ops_abs)} operações | {len(ind_abs)} individuais")

        # Média dos últimos 3 meses
        sorted_meses = sorted(abs_data['geral_por_mes'].keys(), reverse=True)
        ultimos_3 = sorted_meses[:3]
        if ultimos_3:
            abs_data['geral_ultimos_3'] = round(
                sum(abs_data['geral_por_mes'][m] for m in ultimos_3) / len(ultimos_3), 2
            )
            print(f"  ABS Média últimos 3 meses ({ultimos_3}): {abs_data['geral_ultimos_3']}%")
    except Exception as e:
        print(f"  Erro ao ler ABS: {e}")
        traceback.print_exc()

    # ── Salva ─────────────────────────────────────────────────────────
    
    # ── 6. CÁLCULO DINÂMICO DE QUARTIS E DISPERSÃO ────────────────────
    def calcular_quartis(meta_data, adm_list):
        import numpy as np
        
        # Mapear operações
        op_map = {a['nome_norm']: a['operacao'] for a in adm_list if a.get('operacao')}
        
        quartil_result = {}
        
        for mes_key, mes_info in meta_data.items():
            rows = mes_info.get('rows', [])
            
            # Separar por operação
            ops = {}
            for r in rows:
                if not r.get('nome_norm'): continue
                op = op_map.get(r['nome_norm'], 'Desconhecida')
                if op not in ops:
                    ops[op] = []
                ops[op].append(r)
                
            resultados_mes = []
            carteiras_list = []
            
            # Para médias gerais
            all_valid_ho = []
            all_valid_prom = []
            
            for op_name, op_rows in ops.items():
                # Validos = ho > 0
                validos = [r for r in op_rows if r.get('ho', 0) > 0]
                
                ho_n = len(validos)
                if ho_n == 0: continue
                
                # ho dispersão/quartil
                validos.sort(key=lambda x: x.get('ho', 0), reverse=True)
                ho_max = max(x.get('ho', 0) for x in validos)
                
                for rank, r in enumerate(validos):
                    pct = rank / max(1, ho_n - 1)
                    if ho_n <= 1: q = 1
                    elif pct <= 0.25: q = 1
                    elif pct <= 0.50: q = 2
                    elif pct <= 0.75: q = 3
                    else: q = 4
                    
                    r['_q_ho'] = f"{q}º Quartil"
                    r['_disp_ho'] = round((r.get('ho', 0) / ho_max * 100), 1) if ho_max > 0 else 0
                    all_valid_ho.append(r.get('ho', 0))
                    
                # promessas dispersão/quartil
                validos_prom = [r for r in validos if r.get('promessas') is not None]
                prom_n = len(validos_prom)
                validos_prom.sort(key=lambda x: x.get('promessas', 0), reverse=True)
                prom_max = max((x.get('promessas', 0) for x in validos_prom), default=0)
                
                for rank, r in enumerate(validos_prom):
                    pct = rank / max(1, prom_n - 1)
                    if prom_n <= 1: q = 1
                    elif pct <= 0.25: q = 1
                    elif pct <= 0.50: q = 2
                    elif pct <= 0.75: q = 3
                    else: q = 4
                    
                    r['_q_prom'] = f"{q}º Quartil"
                    r['_disp_prom'] = round((r.get('promessas', 0) / prom_max * 100), 1) if prom_max > 0 else 0
                    all_valid_prom.append(r.get('promessas', 0))
                
                # Resumo da carteira (Operação)
                def media(lst): return round(float(np.mean(lst)), 2) if lst else 0.0
                
                def get_vals(q_val, key, disp=True):
                    if disp: return [x.get(f'_disp_{key}') for x in validos if x.get(f'_q_{key}') == q_val and f'_disp_{key}' in x]
                    else: return [x.get(key) for x in validos if x.get(f'_q_{key}') == q_val and key in x]
                
                ho_disp = {f"{q}º Quartil": media(get_vals(f"{q}º Quartil", "ho", True)) for q in range(1,5)}
                ho_prod = {f"{q}º Quartil": media(get_vals(f"{q}º Quartil", "ho", False)) for q in range(1,5)}
                
                prom_disp = {f"{q}º Quartil": media(get_vals(f"{q}º Quartil", "prom", True)) for q in range(1,5)}
                prom_prod = {f"{q}º Quartil": media(get_vals(f"{q}º Quartil", "prom", False)) for q in range(1,5)}

                carteira_stats = {
                    "operacao": op_name,
                    "media_ho": media([x.get('ho') for x in validos]),
                    "media_dispersao_ho": media([x.get('_disp_ho') for x in validos if '_disp_ho' in x]),
                    "ho_quartil_disp": ho_disp,
                    "ho_quartil_prod": ho_prod,
                    "media_promessas": media([x.get('promessas') for x in validos_prom]),
                    "media_dispersao_promessas": media([x.get('_disp_prom') for x in validos_prom if '_disp_prom' in x]),
                    "prom_quartil_disp": prom_disp,
                    "prom_quartil_prod": prom_prod,
                    "qtd_agentes": ho_n
                }
                carteiras_list.append(carteira_stats)
                
                for r in validos:
                    # Injetar os valores direto em r
                    r['quartil_ho'] = r.get('_q_ho')
                    r['dispersao_ho'] = r.get('_disp_ho')
                    r['quartil_promessas'] = r.get('_q_prom')
                    r['dispersao_promessas'] = r.get('_disp_prom')
                    
                    # E também na lista legada
                    resultados_mes.append({
                        "Matricula": r.get('matricula', '—'),
                        "Agente": r.get('nome', '—'),
                        "Operacao": op_name,
                        "HO": r.get('ho'),
                        "Quartil_HO": r.get('_q_ho', '—'),
                        "Dispersao_HO": r.get('_disp_ho'),
                        "Promessas": r.get('promessas'),
                        "Quartil_Promessas": r.get('_q_prom', '—'),
                        "Dispersao_Promessas": r.get('_disp_prom')
                    })
                    
            stats_geral = {
                "total_agentes": len(all_valid_ho),
                "media_ho": media(all_valid_ho),
                "media_promessas": media(all_valid_prom),
                "operacoes": sorted(list(ops.keys())),
                "carteiras": carteiras_list
            }
            
            quartil_result[mes_key] = {
                "data": resultados_mes,
                "stats": stats_geral
            }
            
        return quartil_result

    quartil_data = calcular_quartis(meta_data, adm_list)

    output = {
        'meta':       meta_data,
        'tempos':     tempos_list,
        'adm':        adm_list,
        'compensa':   compensa_map,
        'resumo':     resumo_map,
        'resumo_saldo_final': resumo_saldo_final,
        'resumo_totals': {
            'credito_total_sec': resumo_credito_total,
            'debito_total_sec': resumo_debito_total,
            'credito_individual': resumo_credito_individual,
            'debito_individual': resumo_debito_individual,
        },
        'tempo_logado_media_mensal': tempo_logado_media_mensal,
        'abs_data':   abs_data,
        'operacoes':  sorted(operacoes_set),
        'quartil_data': quartil_data,
        'updated_at': datetime.datetime.now().isoformat()
    }
    js_content = f"var EMBEDDED_DATA = {json.dumps(output, ensure_ascii=False, default=str, indent=4)};\n"
    with open(JS_OUTPUT, 'w', encoding='utf-8') as f:
        f.write(js_content)
    print(f"\n  OK! data_embedded.js gerado com sucesso! ({datetime.datetime.now().strftime('%H:%M:%S')})")

if __name__ == "__main__":
    process_excel()

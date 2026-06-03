import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\TEMPOS 2026.xlsx', data_only=True)
ws = wb['RESUMO']

print('=== ALL RESUMO Data ===')
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    nome = row[0].value
    if not nome or str(nome).strip() in ('', 'nan', 'None'):
        continue
    if 'Atualizado' in str(nome) or 'coluna' in str(nome).lower() or 'Pagamento' in str(nome).lower():
        continue
    credito = row[1].value  # CRÉDITO
    debito = row[2].value   # DÉBITO
    bh = row[3].value       # banco de horas
    deb_final = row[4].value  # Débito final
    saldo = row[5].value    # SALDO FINAL
    bh_ponto = row[6].value  # BH PONTO
    mat = row[7].value      # MATRÍCULA
    operacao = row[8].value  # OPERAÇÃO
    print(f'{nome}:')
    print(f'  CRÉDITO={credito}, DÉBITO={debito}, banco_horas={bh}')
    print(f'  Débito_final={deb_final}, SALDO_FINAL={saldo}')
    print(f'  BH_PONTO={bh_ponto}, MAT={mat}, OP={operacao}')
    print()
